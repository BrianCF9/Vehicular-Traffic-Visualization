from flask import Flask,jsonify,render_template,request,Blueprint
import sqlite3
import json
import uuid
from flask import abort
from flask import Response
import time
from functools import wraps
import jwt
import os
import pandas as pd

from openpyxl import load_workbook

endpoints_services = Blueprint('endpoints_services', __name__)



def get_endpoints():
    return endpoints_services

def requires_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token is missing.'}), 401
        try:
            data = jwt.decode(token, 'secret_key', algorithms=['HS256'])
        except:
            return jsonify({'message': 'Token is invalid.'}), 401
        return f(*args, **kwargs)
    return decorated

services={
    "region":"",
    "filters":{},
    "points": [],
    
}
point={
    "type":"",
    "geoposition":{
        "latitude":"",
        "longitude":"",},
    "body":{
        "commune":"",
        "address":"",
    }
}
filters={
    "school":False,
    "bank":False,
    "market":False
}


data_path="/home/ubuntu/uoct_super_lunes/data/obras"

@endpoints_services.route("/api/uoct/map/Metropolitana/services/obras", methods=["GET"])
def get_obras():
    path=os.path.join(data_path,"obras.xlsx")
    services={
    "region":"Metropolitana",
    "points": []
    }
    

    # Abrir el archivo de Excel en modo de escritura
    workbook = load_workbook(path, read_only=False)

    # Obtener la hoja de cálculo activa
    sheet = workbook.active

    # Obtener el valor de una celda
    celda =sheet.cell(row=1, column=1).value
    
    # recorrer las filas
    
    for row in sheet.iter_rows(min_row=2, min_col=1,max_col=13, max_row=20, values_only=True):
        
        point={ 
            "type":"Obra",
            "geoposition":{
                "latitude":row[4],
                "longitude":row[5],},
            "body":{
                "date_start":row[2],
                "date_end":row[3],
                "commune":row[7],
                "address":row[6],
                "extension":row[8],
                "detail":row[9],
                "status":row[10],
                "schedule":row[11],
                "certainty":row[12],
                
            }
        }
        if point["geoposition"]["latitude"]=="Tramo" or point["geoposition"]["latitude"]=="Tramo ":
            linestring=row[0].replace("LINESTRING ("," ").replace(")","").split(",")

            latitudes=[]
            longitudes=[]
            for i in linestring:
                
                latitudes.append(i.split(" ")[2])
                longitudes.append(i.split(" ")[1])
           
            point["geoposition"]["latitude"]=float(latitudes[int(latitudes.__len__()/2)])
            point["geoposition"]["longitude"]=float(longitudes[int(longitudes.__len__()/2)])
            
        services["points"].append(point)
        
        
    
    
    
    
    
    return jsonify(services)


@endpoints_services.route("/api/uoct/map/Concepcion/services/obras", methods=["GET"])
def get_obras_coquimbo():
    path=os.path.join(data_path,"obras_concepcion.xlsx")
    services={
    "region":"Concepcion",
    "points": []
    }
    

    # Abrir el archivo de Excel en modo de escritura
    workbook = load_workbook(path, read_only=False)

    # Obtener la hoja de cálculo activa
    sheet = workbook.active

    # Obtener el valor de una celda
    celda =sheet.cell(row=1, column=1).value
    
    # recorrer las filas
    
    for row in sheet.iter_rows(min_row=2, min_col=1,max_col=13, max_row=20, values_only=True):
        
        point={ 
            "type":"Obra",
            "geoposition":{
                "latitude":row[4],
                "longitude":row[5],},
            "body":{
                "date_start":row[2],
                "date_end":row[3],
                "commune":row[7],
                "address":row[6],
                "extension":row[8],
                "detail":row[9],
                "status":row[10],
                "schedule":row[11],
                "certainty":row[12],
                
            }
        }
        if point["geoposition"]["latitude"]=="Tramo" or point["geoposition"]["latitude"]=="Tramo ":
            linestring=row[0].replace("LINESTRING ("," ").replace(")","").split(",")

            latitudes=[]
            longitudes=[]
            for i in linestring:
                
                latitudes.append(i.split(" ")[2])
                longitudes.append(i.split(" ")[1])
           
            point["geoposition"]["latitude"]=float(latitudes[int(latitudes.__len__()/2)])
            point["geoposition"]["longitude"]=float(longitudes[int(longitudes.__len__()/2)])
            
        services["points"].append(point)
        
        
    
    
    
    
    
    return jsonify(services)







