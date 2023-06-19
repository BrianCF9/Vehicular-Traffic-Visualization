from flask import Flask,jsonify,render_template,request,Blueprint
import sqlite3
import json
import uuid
from flask import abort
from flask import Response
import time
from functools import wraps
import jwt
import csv
import pandas as pd
import redis
import openpyxl

endpoints_modal = Blueprint('endpoints_modal', __name__)
def get_endpoints():
    return endpoints_modal

def requires_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token is missing.'}), 401
        try:
            data = jwt.decode(token, 'secret_key', algorithms=['HS256'])
        except:
            return jsonify({'message': 'Token is invalid.'}), 401
        return f(*args, **kwargs)
    return decorated

# Función para seleccionar los puntos intermedios de la lista de coordenadas

data_path="/home/ubuntu/uoct_super_lunes/data/datos_rve_resumen"
r = redis.Redis(host='localhost', port=6379, db=0)


@endpoints_modal.route("/api/uoct/map/<string:region>/modal/", methods=["GET"])
def get_congestions(region):
    # Load the workbook
    workbook = openpyxl.load_workbook(data_path + "/datos_rve_resumen.xlsx")
    # Select a worksheet
    if region == "Metropolitana":
        title="Región Metropolitana"
        worksheet = workbook['Región Metropolitana']
    else:
        return jsonify({
                "title": "Región no encontrada",
                "linecharts": [{
                        "title": "",
                        "subtitle": "Km/h",
                        "charts": []
                    }],
            })

    fechas = []
    fechas.append("2_3_2023")
    fechas.append("7_3_2022")
    fechas.append("9_11_2022")
    fechas.append("27_2_2023")
    fechas.append("14_12_2022")
    fechas.append("30_11_2022")
    fechas.append("3_3_2023")
    fechas.append("Super Lunes 2023 - 06/03/2023")
    fechas.append("Super Lunes 2022 - 07/03/2022")
    fechas.append("Último Lunes - 27/02/2023")
    

    
    
            
    # Access a cell value
    #cell_value = worksheet.cell(row=1, column=1).value
    #print(cell_value)
    '''
    response={
                "title": "",
                "linecharts": [],
            }
    linechart = {
                        "title": "",
                        "subtitle": "",
                        "charts": []
                    }
    chart={
                                "item": "",
                                "axis_config": {
                                    "vertical_axis_config": {
                                        "min": "",
                                        "max": "",
                                        "labels_count": 5
                                    },
                                    "horizontal_axis_config": {
                                        "min": "",
                                        "max": "",
                                        "labels_count": "",
                                        "custom_labels": []
                                    }
                                },
                                "data": {
                                    "data_sets": [
                                    ]
                                }
                            }
    custom_label = {
        ""
    }
    dataset ={
                "title": "",
                "entries": []
                        }
    
    entrie={
            "position": 0,
            "value": ""
            }
    '''
    # Iterate over rows
    response={
                "title": title,
                "linecharts": [{
                        "title": "",
                        "subtitle": "Km/h",
                        "charts": []
                    }],
            }
    
    
    
    axis_config1= {
                                "vertical_axis_config": {
                                    "min": "",
                                    "max": "",
                                    "labels_count": 5
                                },
                                "horizontal_axis_config": {
                                    "min": "",
                                    "max": "",
                                    "labels_count": "",
                                    "custom_labels": []
                                }
                            }

    custom_labels1=[]
    min1=0
    max1=0

    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=286):
        for i in range(82, 286, 4):
            custom_labels1.append(row[i].value.replace("velocidad_",""))

    

    
    for row in worksheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=286):
        if row[2].value in fechas:
            if row[2].value != "Transporte público":
                for i in range(82, 286, 4):
                    try:
                        if row[i].value is not None:
                            min1 = float(row[i].value)
                            max1 = float(row[i].value)
                    except:
                        None


        
    
    for row in worksheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=286):
        if row[2].value in fechas:
            if row[2].value != "Transporte público":
                for i in range(82, 286, 4):
                    if row[i].value is not None:
                        try:
                            if float(row[i].value) < min1:
                                min1 = float(row[i].value)
                            if float(row[i].value) > max1:
                                max1 = float(row[i].value)
                        except:
                            None


    
    #axis_config1["vertical_axis_config"]["min"] = max1*-1
    #axis_config1["vertical_axis_config"]["max"] = min1*-1
    axis_config1["vertical_axis_config"]["min"] = min1
    axis_config1["vertical_axis_config"]["max"] = max1
    
    
    axis_config1["horizontal_axis_config"]["custom_labels"] = custom_labels1
    axis_config1["horizontal_axis_config"]["labels_count"] = len(custom_labels1)
    axis_config1["horizontal_axis_config"]["min"] = 0
    axis_config1["horizontal_axis_config"]["max"] = len(custom_labels1)-1

    
    chart1={
                                "item": "Resumen Particulares",
                                "axis_config": axis_config1,
                                "data": {
                                    "data_sets": [
                                    ]
                                }
                            }

            
    for row in worksheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=286):
        if row[2].value in fechas:
            if row[2].value != "Transporte público":
                dataset = {
                    "title": row[2].value,
                    "entries": []
                }
                last_value = 0
                for i in range(82, 286, 4):
                    if row[i].value is not None and row[i].value != "":
                        entrie = {
                            "position": int((i - 82) // 2/2),
                            "value": float(row[i].value) 
                        }
                        last_value = float(row[i].value)
                        dataset["entries"].append(entrie)
                    else:
                        entrie = {
                            "position": int((i - 82) // 2/2),
                            "value": last_value
                        }
                        
                chart1["data"]["data_sets"].append(dataset)


    
    response["linecharts"][0]["charts"].append(chart1)

    
    return jsonify(response)

